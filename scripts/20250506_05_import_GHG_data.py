# 20250430
# -*- coding: utf-8 -*-
import json
import openpyxl
import pandas as pd
import numpy as np

YEAR_START = 1990
YEAR_END   = 2024
year_range = np.arange(YEAR_START, YEAR_END)

# output file names
jsonfile_data = 'outputs/20250506_GHGI/20250506_05_ghg_data.json'
excelfile_data = 'outputs/20250506_GHGI/20250506_05_ghg_data.xlsx'

data_structure_excel = 'inputs/GHGI/20250506_05_data_structure.xlsx'
wb = openpyxl.load_workbook(data_structure_excel)
sheet = wb['Sheet1']
data = sheet.values
cols = next(data)
data = list(data)
df_data_structure = pd.DataFrame(data, columns=cols)
#df1 = df1_0.dropna()
#df1['member_id'] = ''
#df1.reset_index(drop=True, inplace=True)
n_data_structure = df_data_structure.index.shape[0]

ghgi_excel = 'inputs/GHGI/L5-7gas_2025_gioweb_1.0_mod.xlsx'
wb = openpyxl.load_workbook(ghgi_excel, data_only=True)
ghgi_sheet2 = wb['2.CO2-sector']
ghgi_sheet3 = wb['3.Allocated_CO2-sector']
ghgi_sheet6 = wb['6.CH4']
ghgi_sheet7 = wb['7.N2O']
ghgi_sheet8 = wb['8.F-gas']

df1 = pd.DataFrame(
    {
        "id": df_data_structure['id'],
        "label": df_data_structure['label'],
        "level": df_data_structure['level'],
        "l0": df_data_structure['l0'],
        "l1": df_data_structure['l1'],
        "l2": df_data_structure['l2'],
        "l3": df_data_structure['l3'],
        "l4": df_data_structure['l4'],
        "l5": df_data_structure['l5'],
        'n_sub': df_data_structure['n_sub'],
        'unit': df_data_structure['unit'],
        "item_name_jp": '',
        "item_name_en": '',
    }
)
n_df1 = df1.index.shape[0]


## set item names
for i in range(n_df1):
    if df1.loc[i,'l0'] == 1:
        ghgi_sheet = ghgi_sheet2
    elif df1.loc[i,'l0'] == 2:
        ghgi_sheet = ghgi_sheet3
    elif df1.loc[i,'l0'] == 3:
        ghgi_sheet = ghgi_sheet6
    elif df1.loc[i,'l0'] == 4:
        ghgi_sheet = ghgi_sheet7
    elif df1.loc[i,'l0'] == 5:
        ghgi_sheet = ghgi_sheet8
    else:
        print('not supported')
        exit()

    if df1.loc[i,'level'] > 0:
        row = df_data_structure.loc[i,'row']
        col_jp = df_data_structure.loc[i,'col_name_jp']
        col_en = df_data_structure.loc[i,'col_name_en']
        cellname = '%s%d' % (col_jp,row)
        name_jp = ghgi_sheet[cellname].value
        cellname = '%s%d' % (col_en,row)
        name_en = ghgi_sheet[cellname].value
        df1.loc[i,'item_name_jp'] = name_jp
        df1.loc[i,'item_name_en'] = name_en
    else:
        df1.loc[i,'item_name_jp'] = df_data_structure.loc[i,'name_jp']
        df1.loc[i,'item_name_en'] = df_data_structure.loc[i,'name_en']

df2 = pd.DataFrame(columns=year_range)

## import values of those with n_sub = 0
for i in range(n_df1):
    if df1.loc[i,'l0'] == 1:
        ghgi_sheet = ghgi_sheet2
    elif df1.loc[i,'l0'] == 2:
        ghgi_sheet = ghgi_sheet3
    elif df1.loc[i,'l0'] == 3:
        ghgi_sheet = ghgi_sheet6
    elif df1.loc[i,'l0'] == 4:
        ghgi_sheet = ghgi_sheet7
    elif df1.loc[i,'l0'] == 5:
        ghgi_sheet = ghgi_sheet8
    else:
        print('not supported')
        exit()

    if df_data_structure.loc[i,'n_sub'] == 0:
        for j in range(YEAR_END-YEAR_START):
            colpos = 27+j
            v = ghgi_sheet.cell(column=colpos,row=df_data_structure.loc[i,'row']).value
            if v == 'NO':
                v = None
            df2.loc[i,YEAR_START+j]=v


## Level 4
for i in range(n_df1):
    if df1.loc[i,'level'] == 4 and df1.loc[i,'n_sub'] > 0:
        n_sub = df1.loc[i,'n_sub']
        l0 = int(df1.loc[i,'l0'])
        l1 = int(df1.loc[i,'l1'])
        l2 = int(df1.loc[i,'l2'])
        l3 = int(df1.loc[i,'l3'])
        l4 = int(df1.loc[i,'l4'])
        sum = np.zeros(YEAR_END-YEAR_START)
        for j in range(i+1,n_df1):
            if df1.loc[j,'level'] == 5 and df1.loc[j,'l0'] == l0 and df1.loc[j,'l1'] == l1 and df1.loc[j,'l2'] == l2 and df1.loc[j,'l3'] == l3 and df1.loc[j,'l4'] == l4 and df1.loc[j,'l5'] > 0:
                for k in range(YEAR_START, YEAR_END):
                    v = df2.loc[j,k]
                    sum[k-YEAR_START] += v
        for k in range(YEAR_START,YEAR_END):
            df2.loc[i,k] = sum[k-YEAR_START]

## Level 3
for i in range(n_df1):
    if df1.loc[i,'level'] == 3 and df1.loc[i,'n_sub'] > 0:
        n_sub = df1.loc[i,'n_sub']
        l0 = int(df1.loc[i,'l0'])
        l1 = int(df1.loc[i,'l1'])
        l2 = int(df1.loc[i,'l2'])
        l3 = int(df1.loc[i,'l3'])
        sum = np.zeros(YEAR_END-YEAR_START)
        for j in range(i+1,n_df1):
            if df1.loc[j,'level'] == 4 and df1.loc[j,'l0'] == l0 and df1.loc[j,'l1'] == l1 and df1.loc[j,'l2'] == l2 and df1.loc[j,'l3'] == l3 and df1.loc[j,'l4'] > 0:
                for k in range(YEAR_START, YEAR_END):
                    if pd.isnull(df2.loc[j,k]) == False:
                        v = df2.loc[j,k]
                        sum[k-YEAR_START] += v
        for k in range(YEAR_START,YEAR_END):
            df2.loc[i,k] = sum[k-YEAR_START]

## Level 2
for i in range(n_df1):
    if df1.loc[i,'level'] == 2 and df1.loc[i,'n_sub'] > 0:
        n_sub = int(df1.loc[i,'n_sub'])
        l0 = int(df1.loc[i,'l0'])
        l1 = int(df1.loc[i,'l1'])
        l2 = int(df1.loc[i,'l2'])
        sum = np.zeros(YEAR_END-YEAR_START)
        for j in range(i+1,n_df1):
            if df1.loc[j,'level'] == 3 and df1.loc[j,'l0'] == l0 and df1.loc[j,'l1'] == l1 and df1.loc[j,'l2'] == l2 and df1.loc[j,'l3'] > 0:
                for k in range(YEAR_START, YEAR_END):
                    if pd.isnull(df2.loc[j,k]) == False:
                        v = df2.loc[j,k]
                        sum[k-YEAR_START] += v
        for k in range(YEAR_START,YEAR_END):
            df2.loc[i,k] = sum[k-YEAR_START]

## Level 1
for i in range(n_df1):
    if df1.loc[i,'level'] == 1 and df1.loc[i,'n_sub'] > 0:
        n_sub = df1.loc[i,'n_sub']
        l0 = int(df1.loc[i,'l0'])
        l1 = int(df1.loc[i,'l1'])
        sum = np.zeros(YEAR_END-YEAR_START)
        for j in range(i+1,n_df1):
            if df1.loc[j,'level'] == 2 and df1.loc[j,'l0'] == l0 and df1.loc[j,'l1'] == l1 and df1.loc[j,'l2'] > 0:
                for k in range(YEAR_START, YEAR_END):
                    if pd.isnull(df2.loc[j,k]) == False:
                        v = df2.loc[j,k]
                        sum[k-YEAR_START] += v
        for k in range(YEAR_START,YEAR_END):
            df2.loc[i,k] = sum[k-YEAR_START]

## Level 0
for i in range(n_df1):
    if df1.loc[i,'level'] == 0 and df1.loc[i,'n_sub'] > 0:
        n_sub = df1.loc[i,'n_sub']
        l0 = int(df1.loc[i,'l0'])
        sum = np.zeros(YEAR_END-YEAR_START)
        for j in range(i+1,n_df1):
            if df1.loc[j,'level'] == 1 and df1.loc[j,'l0'] == l0 and df1.loc[j,'l1'] > 0:
                for k in range(YEAR_START, YEAR_END):
                    v = df2.loc[j,k]
                    sum[k-YEAR_START] += v
        for k in range(YEAR_START,YEAR_END):
            df2.loc[i,k] = sum[k-YEAR_START]


df3 = pd.concat([df1, df2], axis=1)
#with open(jsonfile_data_structure, 'w', encoding='utf-8') as f:
#    json.dump(df1, f, ensure_ascii=False, indent=4)
df3.to_json(jsonfile_data, orient='index', force_ascii=False, indent=4)
df3.to_excel(excelfile_data)
